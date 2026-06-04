"""Excel .xlsx parser based on openpyxl read-only mode."""

from __future__ import annotations

import io

from fish_worker.parser.base import BaseParser, RawElement


class XlsxParser(BaseParser):
    """Extract non-empty spreadsheet rows as table-like text elements."""

    def parse(self, content: bytes, filename: str, tmp_dir: str | None = None) -> list[RawElement]:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        elements: list[RawElement] = []
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    cells = []
                    for cell in row:
                        if cell is None:
                            continue
                        text = str(cell).strip()
                        if text:
                            cells.append(text)
                    if cells:
                        elements.append(RawElement(text=" | ".join(cells), page=None, element_type="Table"))
        finally:
            workbook.close()
        return elements
